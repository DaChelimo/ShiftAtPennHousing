#!/usr/bin/env python3
"""Generate supabase/seeds/harnwell-summer-sandbox.sql from the real Summer 2026
availability forms + the final schedule Ornella/Purity/Abraham fall back on."""

import openpyxl, os, sys

DL = "/Users/DaChelimo/Downloads/"
OUT = "/Users/DaChelimo/Documents/TechWork/Shift@PennHousing/supabase/seeds/harnwell-summer-sandbox.sql"

# ---------------------------------------------------------------- slot model
# 37 half-hour blocks per day: 05:30 then hourly rows split into two blocks.
ROW_SPANS = [
    ("5:30AM - 6:00AM", 330, 360),
    ("6:00AM - 7:00AM", 360, 420),
    ("7:00AM - 8:00AM", 420, 480),
    ("8:00AM - 9:00AM", 480, 540),
    ("9:00AM - 10:00AM", 540, 600),
    ("10:00AM - 11:00AM", 600, 660),
    ("11:00AM - Noon", 660, 720),
    ("Noon - 1:00PM", 720, 780),
    ("1:00PM - 2:00PM", 780, 840),
    ("2:00PM - 3:00PM", 840, 900),
    ("3:00PM - 4:00PM", 900, 960),
    ("4:00PM - 5:00PM", 960, 1020),
    ("5:00PM - 6:00PM", 1020, 1080),
    ("6:00PM - 7:00PM", 1080, 1140),
    ("7:00PM - 8:00PM", 1140, 1200),
    ("8:00PM - 9:00PM", 1200, 1260),
    ("9:00PM - 10:00PM", 1260, 1320),
    ("10:00PM - 11:00PM", 1320, 1380),
    ("11:00PM - Midnight", 1380, 1440),
]
MINUTES = [330] + [m for h in range(6, 24) for m in (h * 60, h * 60 + 30)]
assert len(MINUTES) == 37

DAYS = list(range(7))  # 0=Mon .. 6=Sun


def blank():
    """weekday -> {minute: status}, defaulting to 'available' (the painter's neutral)."""
    return {d: {m: "available" for m in MINUTES} for d in DAYS}


def paint(grid, days, lo, hi, status):
    for d in days:
        for m in MINUTES:
            if lo <= m < hi:
                grid[d][m] = status


# ---------------------------------------------------------------- xlsx readers
STATUS = {"preferred": "preferred", "available": "available", "cannot": "cannot"}


def read_text_grid(path):
    ws = openpyxl.load_workbook(DL + path, data_only=True)["Sheet1"]
    grid = blank()
    for i, (_label, lo, hi) in enumerate(ROW_SPANS):
        r = 25 + i
        for d, col in enumerate("BCDEFGH"):
            v = ws[f"{col}{r}"].value
            if v is None:
                continue
            s = STATUS[str(v).strip().lower()]
            paint(grid, [d], lo, hi, s)
    return grid


def read_color_grid(path, legend):
    ws = openpyxl.load_workbook(DL + path)["Sheet1"]
    grid = blank()
    for i, (_label, lo, hi) in enumerate(ROW_SPANS):
        r = 25 + i
        for d, col in enumerate("BCDEFGH"):
            c = ws[f"{col}{r}"]
            rgb = c.fill.fgColor.rgb if c.fill and c.fill.fgColor else None
            if not isinstance(rgb, str) or rgb not in legend:
                continue
            paint(grid, [d], lo, hi, legend[rgb])
    return grid


# ---------------------------------------------------------------- final schedule
def final_schedule_slots():
    """name -> set of (weekday, minute) they actually worked."""
    ws = openpyxl.load_workbook(DL + "Final Schedule.xlsx", data_only=True)["Summer Final"]
    # two seat columns per day, Mon..Sun
    cols = [("B", "C"), ("D", "E"), ("F", "G"), ("H", "I"), ("J", "K"), ("L", "M"), ("N", "O")]
    worked = {}
    for i, (_label, lo, hi) in enumerate(ROW_SPANS):
        r = 2 + i
        for d, pair in enumerate(cols):
            for col in pair:
                v = ws[f"{col}{r}"].value
                if v is None:
                    continue
                nm = str(v).strip()
                if nm.upper() == "OPEN":
                    continue
                for m in MINUTES:
                    if lo <= m < hi:
                        worked.setdefault(nm, set()).add((d, m))
    return worked


WORKED = final_schedule_slots()


def from_final(name, rest_status):
    grid = blank()
    for d in DAYS:
        for m in MINUTES:
            grid[d][m] = rest_status
    for (d, m) in WORKED[name]:
        grid[d][m] = "preferred"
    return grid


# ---------------------------------------------------------------- the nine
def valeria():
    # Valeria.pdf: Mon-Fri only; Sat/Sun entirely Cannot.
    g = blank()
    paint(g, [0, 1, 2, 3, 4], 330, 480, "cannot")       # 5:30-8:00
    paint(g, [0, 1, 2, 3, 4], 480, 720, "preferred")    # 8:00-12:00
    paint(g, [0, 1, 2, 3, 4], 720, 1020, "available")   # 12:00-17:00
    paint(g, [0, 1, 2, 3, 4], 1020, 1260, "preferred")  # 17:00-21:00
    paint(g, [0, 1, 2, 3, 4], 1260, 1440, "available")  # 21:00-24:00
    paint(g, [5, 6], 330, 1440, "cannot")
    return g


def abraham():
    # No form. Stakeholder rule: 05:30-08:00 Cannot every day, everything else Preferred.
    g = blank()
    paint(g, DAYS, 330, 1440, "preferred")
    paint(g, DAYS, 330, 480, "cannot")
    return g


WORKERS = [
    # (db name, email, target_hours, grid, provenance)
    ("Eleni", "elenikan@sas.upenn.edu", 23,
     read_color_grid("ELENI.xlsx", {"FF00B050": "preferred", "FFFFFF00": "available", "FFFF0000": "cannot"}),
     "ELENI.xlsx (colour-coded form; hours from the final schedule contact table)"),
    ("Abraham", "ndlovuab@sas.upenn.edu", 24, abraham(),
     "no form: stakeholder rule 05:30-08:00 Cannot daily, else Preferred"),
    ("Drew", "dbukasa@sas.upenn.edu", 30,
     read_color_grid("ANDREW BUKASA.xlsx", {"FF339966": "preferred", "FFFF9900": "available", "FFFF0000": "cannot"}),
     "ANDREW BUKASA.xlsx (colour-coded form; stated 23-30, upper bound used)"),
    ("Valeria", "mercadov@sas.upenn.edu", 30, valeria(), "Valeria.pdf"),
    ("Aaron", "akkirui@sas.upenn.edu", 23, read_text_grid("Aaron Kirui.xlsx"), "Aaron Kirui.xlsx"),
    ("Lealem", "lmelesse@seas.upenn.edu", 30, read_text_grid("Lealem.xlsx"), "Lealem.xlsx"),
    ("Ornella", "ornellar@sas.upenn.edu", 24, from_final("Ornella", "cannot"),
     "no form: derived from the final schedule (worked = Preferred, all else Cannot)"),
    ("Andrew Chelimo", "chelimo@seas.upenn.edu", 40, read_text_grid("Andrew Chelimo .xlsx"),
     "Andrew Chelimo .xlsx"),
    ("Purity", "liseche1@nursing.upenn.edu", 40, read_text_grid("Purity.xlsx"), "Purity.xlsx"),
]

# ---------------------------------------------------------------- emit
def compress(grid):
    """weekday -> [(lo, hi, status)] run-length ranges over MINUTES."""
    out = {}
    for d in DAYS:
        runs = []
        for m in MINUTES:
            s = grid[d][m]
            end = m + 30
            if runs and runs[-1][2] == s and runs[-1][1] == m:
                runs[-1] = (runs[-1][0], end, s)
            else:
                runs.append((m, end, s))
        out[d] = runs
    return out


def hhmm(m):
    return f"{m // 60:02d}:{m % 60:02d}"


DAYNAME = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

rows = []
for name, email, target, grid, prov in WORKERS:
    rows.append((name, email, target, compress(grid), prov))

lines = []
w = lines.append
w("-- supabase/seeds/harnwell-summer-sandbox.sql")
w("-- AUTO-GENERATED. Regenerate with `pnpm seed:sandbox:regen` (scripts/gen-harnwell-sandbox.py).")
w("--")
w("-- SANDBOX: the nine REAL Harnwell Summer 2026 student workers, with the preferences")
w("-- they actually submitted on the paper availability forms, written against the")
w("-- Summer 2026 scheduling period's template week (Mon 2026-06-01 .. Sun 2026-06-07).")
w("-- Purpose: drive the AI schedule builder from real inputs and compare its output to")
w("-- the schedule the student manager actually published ('Final Schedule.xlsx').")
w("--")
w("-- Containment: this touches ONLY (a) preferences + period_targets rows belonging to")
w("-- Harnwell-home users in the Summer 2026 period, and (b) Harnwell draft assignments in")
w("-- that period. It never touches other houses, other periods, live shift assignments,")
w("-- users, or auth. Re-running it fully resets the sandbox to this exact state, which is")
w("-- what makes it safe to iterate on the AI agent and re-test.")
w("--")
w("-- The eight synthetic Harnwell workers (Alice/Ben/Cara/Dan/Erin/Fred/Gina/Hugo) keep")
w("-- their accounts but lose their simulated preferences for this period, which drops them")
w("-- from the builder roster: getAiScheduleContext keeps SUBMITTERS ONLY (>= 1 preference")
w("-- row or a period_targets row). That is how the sandbox isolates the nine real workers")
w("-- without deactivating anybody. Re-run `Simulate worker preferences` on /admin/operations")
w("-- to put the synthetic cast back.")
w("--")
w("-- Sources (one row per worker, provenance recorded):")
for name, _e, target, _r, prov in rows:
    w(f"--   {name:<15} target {target:>2}h  <- {prov}")
w("")
w("BEGIN;")
w("")
w("-- 1. Guard: the sandbox is meaningless without its period and its template week.")
w("DO $$")
w("BEGIN")
w("  IF NOT EXISTS (SELECT 1 FROM scheduling_periods WHERE period_id = '5ea50000-0000-4000-8000-000000000001') THEN")
w("    RAISE EXCEPTION 'Summer 2026 period missing. Run `pnpm db:reset:seasons` first.';")
w("  END IF;")
w("  IF NOT EXISTS (SELECT 1 FROM users WHERE user_id = 'fbb00000-0000-4000-8000-000000000001') THEN")
w("    RAISE EXCEPTION 'Real Harnwell workers missing. Run `pnpm seed:harnwell` first.';")
w("  END IF;")
w("END $$;")
w("")
w("-- 2. Abraham is the Harnwell SM and also works the desk (24h on the real schedule),")
w("--    but house_roster_as_of is role = 'sw' only, so without this he can never appear")
w("--    in the builder or AI roster. Grant him the sw role too (roles are additive).")
w("INSERT INTO user_roles (user_id, role, scope_house_id)")
w("SELECT user_id, 'sw', NULL FROM users WHERE email = 'ndlovuab@sas.upenn.edu'")
w("ON CONFLICT DO NOTHING;")
w("")
w("-- 3. The submitted grids, as half-hour ranges over the template week.")
w("--    weekday: 0 = Monday .. 6 = Sunday (matches blockWeekSlot).")
w("CREATE TEMP TABLE sandbox_prefs (")
w("  email      text NOT NULL,")
w("  weekday    int  NOT NULL,")
w("  start_min  int  NOT NULL,")
w("  end_min    int  NOT NULL,")
w("  status     preference_status_enum NOT NULL")
w(") ON COMMIT DROP;")
w("")
w("INSERT INTO sandbox_prefs (email, weekday, start_min, end_min, status) VALUES")
tuples = []
for name, email, target, runs, prov in rows:
    tuples.append(f"-- {name}")
    for d in DAYS:
        for (lo, hi, s) in runs[d]:
            tuples.append(
                f"  ('{email}', {d}, {lo}, {hi}, '{s}'),"
                f"  -- {DAYNAME[d]} {hhmm(lo)}-{hhmm(hi % 1440) if hi != 1440 else '24:00'}"
            )
# strip trailing comma on the last data row
last = max(i for i, t in enumerate(tuples) if t.startswith("  ("))
head, sep, tail = tuples[last].partition("),")
tuples[last] = head + ")" + tail
lines.extend(tuples)
w(";")
w("")
w("CREATE TEMP TABLE sandbox_targets (email text NOT NULL, target_hours int NOT NULL) ON COMMIT DROP;")
w("INSERT INTO sandbox_targets (email, target_hours) VALUES")
tg = [f"  ('{e}', {t})," for _n, e, t, _r, _p in rows]
tg[-1] = tg[-1].rstrip(",")
lines.extend(tg)
w(";")
w("")
w("-- 4. The template week's Harnwell blocks, keyed by (weekday, minute-of-day).")
w("CREATE TEMP TABLE sandbox_blocks AS")
w("SELECT b.block_id,")
w("       (EXTRACT(isodow FROM b.block_start_at AT TIME ZONE 'America/New_York')::int - 1) AS weekday,")
w("       (EXTRACT(hour FROM b.block_start_at AT TIME ZONE 'America/New_York')::int * 60")
w("        + EXTRACT(minute FROM b.block_start_at AT TIME ZONE 'America/New_York')::int) AS minute_of_day")
w("FROM shift_blocks b")
w("WHERE b.house_id = 'harnwell'")
w("  AND b.voided_at IS NULL")
w("  AND b.block_start_at >= TIMESTAMPTZ '2026-06-01 00:00:00-04'")
w("  AND b.block_start_at <  TIMESTAMPTZ '2026-06-08 00:00:00-04';")
w("")
w("DO $$")
w("DECLARE v_n int;")
w("BEGIN")
w("  SELECT count(*) INTO v_n FROM sandbox_blocks;")
w("  IF v_n <> 259 THEN")
w("    RAISE EXCEPTION 'Expected 259 Harnwell template-week blocks (37 x 7), found %.', v_n;")
w("  END IF;")
w("END $$;")
w("")
w("-- 5. Open the preference window for the write. enforce_preference_deadline fires on")
w("--    INSERT/UPDATE *and DELETE* and is not service-role-bypassed, so the delete below")
w("--    must happen inside the reopened window too. NULL deadline = open. Restored in 9.")
w("CREATE TEMP TABLE sandbox_saved_deadline ON COMMIT DROP AS")
w("SELECT preference_deadline FROM scheduling_periods")
w("WHERE period_id = '5ea50000-0000-4000-8000-000000000001' FOR UPDATE;")
w("")
w("UPDATE scheduling_periods SET preference_deadline = NULL")
w("WHERE period_id = '5ea50000-0000-4000-8000-000000000001';")
w("")
w("-- 6. Clear every Harnwell-home user's preferences/targets for this period (real AND")
w("--    synthetic), so the sandbox is the only thing left standing. Other houses untouched.")
w("DELETE FROM preferences p")
w("USING users u")
w("WHERE u.user_id = p.user_id")
w("  AND u.home_house_id = 'harnwell'")
w("  AND p.period_id = '5ea50000-0000-4000-8000-000000000001';")
w("")
w("DELETE FROM period_targets t")
w("USING users u")
w("WHERE u.user_id = t.user_id")
w("  AND u.home_house_id = 'harnwell'")
w("  AND t.period_id = '5ea50000-0000-4000-8000-000000000001';")
w("")
w("-- 7. Write the nine grids. Every block gets an explicit status, exactly like the")
w("--    in-app painter's full-grid upsert (buildSubmitPayload).")
w("INSERT INTO preferences (user_id, block_id, period_id, status)")
w("SELECT u.user_id, sb.block_id, '5ea50000-0000-4000-8000-000000000001', sp.status")
w("FROM sandbox_prefs sp")
w("JOIN users u ON u.email = sp.email")
w("JOIN sandbox_blocks sb")
w("  ON sb.weekday = sp.weekday")
w(" AND sb.minute_of_day >= sp.start_min")
w(" AND sb.minute_of_day <  sp.end_min")
w("ON CONFLICT (user_id, block_id, period_id) DO UPDATE SET status = EXCLUDED.status;")
w("")
w("INSERT INTO period_targets (user_id, period_id, target_hours, opted_out)")
w("SELECT u.user_id, '5ea50000-0000-4000-8000-000000000001', st.target_hours, false")
w("FROM sandbox_targets st JOIN users u ON u.email = st.email")
w("ON CONFLICT (user_id, period_id) DO UPDATE")
w("  SET target_hours = EXCLUDED.target_hours, opted_out = EXCLUDED.opted_out;")
w("")
w("-- 8. Clear Harnwell drafts for the period so each AI run starts from an empty grid.")
w("DELETE FROM draft_block_assignments d")
w("USING shift_blocks b")
w("WHERE b.block_id = d.block_id")
w("  AND b.house_id = 'harnwell'")
w("  AND d.period_id = '5ea50000-0000-4000-8000-000000000001';")
w("")
w("-- 9. Restore the deadline exactly as found (the sandbox never changes the window).")
w("UPDATE scheduling_periods sp SET preference_deadline = s.preference_deadline")
w("FROM sandbox_saved_deadline s")
w("WHERE sp.period_id = '5ea50000-0000-4000-8000-000000000001';")
w("")
w("-- 10. Verify: 9 workers x 259 blocks = 2331 preference rows, 9 targets, 0 drafts.")
w("DO $$")
w("DECLARE v_p int; v_t int;")
w("BEGIN")
w("  SELECT count(*) INTO v_p FROM preferences p JOIN users u USING (user_id)")
w("   WHERE u.home_house_id = 'harnwell' AND p.period_id = '5ea50000-0000-4000-8000-000000000001';")
w("  SELECT count(*) INTO v_t FROM period_targets t JOIN users u USING (user_id)")
w("   WHERE u.home_house_id = 'harnwell' AND t.period_id = '5ea50000-0000-4000-8000-000000000001';")
w("  IF v_p <> 2331 OR v_t <> 9 THEN")
w("    RAISE EXCEPTION 'Sandbox verification failed: % preference rows (want 2331), % targets (want 9).', v_p, v_t;")
w("  END IF;")
w("  RAISE NOTICE 'Harnwell summer sandbox ready: 9 workers, % preference rows, % targets.', v_p, v_t;")
w("END $$;")
w("")
w("COMMIT;")
w("")

with open(OUT, "w") as f:
    f.write("\n".join(lines))

# ---------------------------------------------------------------- report
print(f"wrote {OUT}")
print()
print(f"{'worker':<16}{'target':>7}{'pref h':>8}{'avail h':>9}{'cannot h':>10}   actual (final sched)")
actual = {n: len(s) / 2 for n, s in WORKED.items()}
alias = {"Andrew Chelimo": "Andrew C."}
for name, email, target, grid, prov in WORKERS:
    c = {"preferred": 0, "available": 0, "cannot": 0}
    for d in DAYS:
        for m in MINUTES:
            c[grid[d][m]] += 1
    a = actual.get(alias.get(name, name), 0)
    print(f"{name:<16}{target:>7}{c['preferred']/2:>8.1f}{c['available']/2:>9.1f}{c['cannot']/2:>10.1f}   {a:>5.1f}h")
print()
print("final-schedule roster:", sorted(actual))
